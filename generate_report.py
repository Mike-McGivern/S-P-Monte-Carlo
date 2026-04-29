import pandas as pd
import seaborn as sns 
import matplotlib.pyplot as plt
import numpy as np
import scipy

import time
import os
import argparse
import requests

from datetime import datetime

import yfinance as yf

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image

parser = argparse.ArgumentParser()
parser.add_argument("--index", action="store_true")
parser.add_argument("--constituents", action="store_true")
parser.add_argument("--volatility", action="store_true")
parser.add_argument("--momentum", action="store_true")
parser.add_argument("--sector", action="store_true")
parser.add_argument("--risk", action="store_true")
parser.add_argument("--liquidity", action="store_true")
parser.add_argument("--breadth", action="store_true")
args = parser.parse_args()

def generate_index_sheet(filepath):
    
    ticker_symbol = "^GSPC"
    ticker = yf.Ticker(ticker_symbol)
    df = ticker.history(period = "1mo") #make period variable in final build

    df.index = df.index.tz_localize(None)

    df['Daily Return'] = df['Close'].pct_change() * 100

    timestamp = datetime.now().strftime("%m-%d-%Y_%H-%M-%S")



    df.to_excel(filepath, sheet_name = "SP500 Data", index = True)

    wb = load_workbook(filepath)
    ws = wb["SP500 Data"]
    ws.move_range("A1:Z9999", rows = 2, cols = 0)
    ws["A1"] = f"Ticker: {ticker_symbol}"
    ws['A2'] = f"Report generated on: {timestamp}"

    for cell in ws[3]:
        cell.font = Font(bold = True)

    daily_return_col = None
    for cell in ws[3]:
        if cell.value == "Daily Return":
            daily_return_col = cell.column_letter
            break
    if daily_return_col:
        ws.conditional_formatting.add(
            f"{daily_return_col}4:{daily_return_col}5000",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
        )
        ws.conditional_formatting.add(
            f"{daily_return_col}4:{daily_return_col}5000",
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )
    
    for column_cell in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cell)
        column_letter = get_column_letter(column_cell[0].column)
        ws.column_dimensions[column_letter].width = length + 2
    
    ws.freeze_panes = "A4"

    wb.save(filepath)

    print("Index data added to:", filepath)
    return df

def generate_constituent_sheet(filepath, global_df):
    print("Preparing SP500 constituent sheet...")

    # 1. Keep only the columns you need
    df = global_df.copy()
    df = df[["Date", "Ticker", "Open", "High", "Low", "Close", "Volume", "Daily Returns"]]

    # 2. Sort for grouping
    df = df.sort_values(["Ticker", "Date"])

    # 3. Flatten the data for Excel
    flat = df.reset_index(drop = True)
    blocks = []
    for ticker, group in flat.groupby("Ticker"):
        blocks.append(group)
        blank = pd.DataFrame([[""] * len(flat.columns)], columns = group.columns)
        blocks.append(blank)
    flat_pretty = pd.concat(blocks, ignore_index = True)

    mode = "a" if os.path.exists(filepath) else "w"

    with pd.ExcelWriter(filepath, engine="openpyxl", mode = mode) as writer:
        flat_pretty.to_excel(writer, sheet_name = "SP500 Constituents", index = False)

    wb = load_workbook(filepath)

    ws = wb["SP500 Constituents"]

    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
        )
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 2
    

    daily_return_col = None
    for cell in ws[1]:
        if cell.value == "Daily Returns":
            daily_return_col = cell.column_letter
            break
    if daily_return_col:
        ws.conditional_formatting.add(
            f"{daily_return_col}2:{daily_return_col}50000",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
        )
        ws.conditional_formatting.add(
            f"{daily_return_col}2:{daily_return_col}50000",
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )
    
    wb.save(filepath)
    print(f"Constituent data added to: {filepath}")
    return flat_pretty

def generate_volatility_sheet(filepath, global_df):
    df = global_df[["Ticker", "Date", "Close", "Daily Returns"]].copy()

    # Pivot for volatility calculations
    close_wide = df.pivot(index="Date", columns="Ticker", values="Close")
    returns_wide = df.pivot(index="Date", columns="Ticker", values="Daily Returns") / 100
    close_prices = close_wide.sort_index()
    daily_returns = close_prices.pct_change()
    rolling_vol = daily_returns.rolling(window=20).std()
    latest_rolling_vol = rolling_vol.iloc[-1] * 100
    vol_df = pd.DataFrame({
        "Ticker": daily_returns.columns,
        "Avg Daily Return (%)": daily_returns.mean() * 100,
        "Daily Volatility (%)": daily_returns.std() * 100,
        "Annualized Volatility (%)": daily_returns.std() * (252 ** 0.5) * 100,
        "Latest Rolling Volatility (%)": latest_rolling_vol,
        "1-Month Return (%)": (close_prices.iloc[-1] / close_prices.iloc[0] - 1) * 100,
    })

    def max_drawdown(series):
        cumulative = (1 + series / 100).cumprod()
        peak = cumulative.cummax()
        drawdown = (cumulative - peak) / peak
        return drawdown.min() * 100
    
    vol_df["Max Drawdown (%)"] = daily_returns.apply(max_drawdown)

    vol_mean = vol_df["Annualized Volatility (%)"].mean()
    vol_std = vol_df["Annualized Volatility (%)"].std()
    vol_df["Volatility Z-Score"] = (vol_df["Annualized Volatility (%)"] - vol_mean) / vol_std

    mode = "a" if os.path.exists(filepath) else "w"
    with pd.ExcelWriter(filepath, engine = "openpyxl", mode = mode) as writer:
        vol_df.to_excel(writer, sheet_name = "SP500 Volatility", index = False)
    
    wb = load_workbook(filepath)
    ws = wb["SP500 Volatility"]

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 2
    
    ws.freeze_panes = "A2"

    vol_col = None
    for cell in ws[1]:
        if cell.value == "Annualized Volatility (%)":
            vol_col = cell.column_letter
            break
    
    if vol_col:
        rng = f"{vol_col}2:{vol_col}{ws.max_row}"
        # Green for low vol
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="lessThan", formula=["30"],
               fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )

        # Yellow for normal vol
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="between", formula=["30", "60"],
               fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"))
        )

        # Red for high vol
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThan", formula=["60"],
                       fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )
        
    z_col = None
    for cell in ws[1]:
        if cell.value == "Volatility Z-Score":
            z_col = cell.column_letter
            break
    if z_col:
        rng = f"{z_col}2:{z_col}{ws.max_row}"

        # Low volatility (green)
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="lessThan", formula=["-1"],
                       fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
        )
        # Normal volatility (yellow)
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="between", formula=["-1", "1"],
                       fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"))
        )
        # High volatility (light red)
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="between", formula=["1", "2"],
                       fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )
        # Extreme volatility (dark red)
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThan", formula=["2"],
                       fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
        )

    wb.save(filepath)
    print(f"Volatility data added to: {filepath}")
    
    return vol_df, daily_returns

def generate_momentum_sheet(filepath, global_df):
    df = global_df[["Ticker", "Date", "Close"]].copy()

    # Pivot to wide format for return windows
    close_wide = df.pivot(index="Date", columns="Ticker", values="Close")

    def compute_returns(prices, periods):
        if len(prices) < periods:
            return pd.Series([None] * prices.shape[1], index = prices.columns)
        returns = (prices.iloc[-1] / prices.iloc[-periods] - 1) * 100
        return returns
    momentum_df = pd.DataFrame({
        "Ticker": close_wide.columns,
        "1-Month Return (%)": compute_returns(close_wide, 21),
        "3-Month Return (%)": compute_returns(close_wide, 63),
        "6-Month Return (%)": compute_returns(close_wide, 126),
    })

    six_month = momentum_df["6-Month Return (%)"]
    momentum_df["Momentum Z-Score"] = (six_month - six_month.mean()) / six_month.std()
    
    momentum_df["Momentum Rank"] = momentum_df["6-Month Return (%)"].rank(ascending = False)    

    mode = "a" if os.path.exists(filepath) else "w"
    with pd.ExcelWriter(filepath, engine = "openpyxl", mode = mode) as writer:
        momentum_df.to_excel(writer, sheet_name = "SP500 Momentum", index = False)
    wb = load_workbook(filepath)
    ws = wb["SP500 Momentum"]

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 2
    
    ws.freeze_panes = "A2"
    z_col = None
    for cell in ws[1]:
        if cell.value == "Momentum Z-Score":
            z_col = cell.column_letter
            break
    if z_col:
        rng = f"{z_col}2:{z_col}{ws.max_row}"

        # Low momentum (red)
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="lessThan", formula=["-1"],
                       fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )
        # Normal momentum (yellow)
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="between", formula=["-1", "1"],
                       fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"))
        )
        # High momentum (green)
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThan", formula=["1"],
                       fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
        )
    wb.save(filepath)
    wb.close()
    print(f"Momentum data added to: {filepath}")

    return momentum_df


def generate_sector_sheet(filepath):
    #sp500 = pd.read_csv("https://raw.githubusercontent.com/datasets/s-and-p-500-companies/master/data/constituents.csv")
    url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
    headers = {"User-Agent": "Mozilla/5.0"}

    html = requests.get(url, headers=headers).text
    sp500 = pd.read_html(html, header=0)[0]

    sp500["Symbol"] = sp500["Symbol"].str.replace(".", "-", regex = False)

    if "GICS Sector" not in sp500.columns:
        print("Warning: Sector column missing — filling with 'Unknown'")
        sp500["Sector"] = "Unknown"

    sp500 = sp500.rename(columns={"GICS Sector": "Sector"})

    sector_map = sp500.set_index("Symbol")["Sector"]
    sector_map = sector_map.str.strip().str.title()

    vol_df = pd.read_excel(filepath, sheet_name = "SP500 Volatility")
    mom_df = pd.read_excel(filepath, sheet_name = "SP500 Momentum")

    if mom_df.empty:
        print("Momentum sheet missing or empty — cannot generate sector sheet yet.")
        return
    if vol_df.empty:
        print("Volatility sheet missing or empty — cannot generate sector sheet yet.")
        return

    vol_df = vol_df.rename(columns={"Symbol": "Ticker"})
    mom_df = mom_df.rename(columns={"Symbol": "Ticker"})

    merged = (
        vol_df.merge(mom_df, on = "Ticker", how = "inner", suffixes=('_vol', '_mom'))
            .merge(sp500[["Symbol", "Sector"]], left_on = "Ticker", right_on = "Symbol", how = "left")
    )

    print("Merged columns:", merged.columns.tolist())

    sector_group = merged.groupby("Sector")
    
    sector_df = pd.DataFrame({
        "Avg 1M Return (%)": sector_group["1-Month Return (%)_mom"].mean(),
        "Avg 3M Return (%)": sector_group["3-Month Return (%)"].mean(),
        "Avg 6M Return (%)": sector_group["6-Month Return (%)"].mean(),
        "Avg Annualized Vol (%)": sector_group["Annualized Volatility (%)"].mean(),
        "Avg Rolling Vol (%)": sector_group["Latest Rolling Volatility (%)"].mean(),
        "Avg Drawdown (%)": sector_group["Max Drawdown (%)"].mean(),
        "Avg Momentum Z-Score": sector_group["Momentum Z-Score"].mean(),
        "Avg Volatility Z-Score": sector_group["Volatility Z-Score"].mean(),
        "Stock Count": sector_group.size(),
    })

    sector_df["Momentum Score"] = (
        sector_df["Avg 1M Return (%)"] * 0.2 +
        sector_df["Avg 3M Return (%)"] * 0.3 +
        sector_df["Avg 6M Return (%)"] * 0.5
    )

    sector_df["Risk Score"] = (
        sector_df["Avg Annualized Vol (%)"] * 0.4 +
        sector_df["Avg Rolling Vol (%)"] * 0.3 +
        (-sector_df["Avg Drawdown (%)"]) * 0.3
    )

    mode = "a" if os.path.exists(filepath) else "w"
    with pd.ExcelWriter(filepath, engine = "openpyxl", mode = mode) as writer:
        sector_df.to_excel(writer, sheet_name = "SP500 Sectors", index = True)
    
    wb = load_workbook(filepath)
    ws = wb["SP500 Sectors"]
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 2

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"Sector data added to: {filepath}")

    return (sector_df, sector_map)

def generate_risk_sheet(filepath, corr, sector_df, vol_df, mom_df):
    wb = load_workbook(filepath)
    if "Risk" in wb.sheetnames:
        ws = wb["Risk"]
        wb.remove(ws)
    ws = wb.create_sheet("Risk")

    # --- CLEAN VOL & MOM ---
    if vol_df.index.name == "Ticker":
        vol_df.index.name = None
        vol_df = vol_df.reset_index()

    if mom_df.index.name == "Ticker":
        mom_df.index.name = None
        mom_df = mom_df.reset_index()

    # --- BUILD STOCK RISK DF ---
    stock_risk_df = vol_df.merge(mom_df, on="Ticker", how="inner")

    stock_risk_df["Risk Score"] = (
        stock_risk_df["Annualized Volatility (%)"] * 0.4 +
        stock_risk_df["Latest Rolling Volatility (%)"] * 0.3 +
        (-stock_risk_df["Max Drawdown (%)"]) * 0.3
    )

    # --- CORRELATION SUMMARY ---
    ws["A1"] = "Correlation Summary"
    ws["A3"] = "Highest Positive Correlation"
    ws["A4"] = "Highest Negative Correlation"
    ws["A5"] = "Average Correlation"

    corr_unstacked = corr.where(~np.eye(corr.shape[0], dtype=bool)).unstack().dropna()
    max_pos = corr_unstacked.idxmax()
    max_neg = corr_unstacked.idxmin()

    ws["B3"] = f"{max_pos[0]} with {max_pos[1]} ({corr_unstacked.max():.4f})"
    ws["B4"] = f"{max_neg[0]} with {max_neg[1]} ({corr_unstacked.min():.4f})"
    ws["B5"] = float(corr_unstacked.mean())


    # Ensure sector name is a column
    if sector_df.index.name is not None:
        sector_df = sector_df.reset_index().rename(columns={"index": "Sector"})

    sector_start = 8
    ws["A7"] = "Sector-Level Risk Summary"

    # Write sector header
    for col_idx, col_name in enumerate(sector_df.columns, start=1):
        ws.cell(row=sector_start, column=col_idx, value=col_name)

    # Write sector rows
    for row_offset, row in enumerate(sector_df.itertuples(index=False), start=1):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=sector_start + row_offset, column=col_idx, value=val)

    stock_start = sector_start + len(sector_df) + 3
    ws.cell(row=stock_start - 1, column=1, value="Stock-Level Risk Summary")

    # Write stock header
    for col_idx, col_name in enumerate(stock_risk_df.columns, start=1):
        ws.cell(row=stock_start, column=col_idx, value=col_name)

    # Write stock rows
    for row_offset, row in enumerate(stock_risk_df.itertuples(index=False), start=1):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=stock_start + row_offset, column=col_idx, value=val)

    # Auto-fit columns
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 2

    ws.freeze_panes = "A23"
    wb.save(filepath)
    print(f"Risk data added to: {filepath}")

    return stock_risk_df

def generate_liquidity_sheet(filepath, global_df, sector_df, sector_map):
    print("Preparing SP500 liquidity sheet...")

    df = global_df[["Ticker", "Date", "Close", "Volume", "Daily Returns"]].copy()

    df["ADV"] = (
        df.groupby("Ticker")["Volume"]
            .transform(lambda x: x.rolling(window=20).mean())
            .reset_index(level=0, drop=True)
    )

    df["DollarVolume"] = df["Close"] * df["Volume"]

    df["ADDV"] = (
        df.groupby("Ticker")["DollarVolume"].transform(lambda x: x.rolling(window=20).mean())
            .reset_index(level=0, drop=True)
    )

    df["Amihud"] = (
        (df["Daily Returns"].abs() / df["DollarVolume"])
            .groupby(df["Ticker"])
            .transform(lambda x: x.rolling(window=20).mean())
            .reset_index(level=0, drop=True)
    )

    latest = df.sort_values("Date").groupby("Ticker").tail(1)

    print("Sector DF columns:", sector_df.columns.tolist())

    latest["Sector"] = latest["Ticker"].map(sector_map)

    latest["Liquidity Rank"] = latest["Amihud"].rank(ascending=True)

    mode = "a" if os.path.exists(filepath) else "w"
    with pd.ExcelWriter(filepath, engine="openpyxl", mode=mode) as writer:
        latest.to_excel(writer, sheet_name="SP500 Liquidity", index=False)
    wb = load_workbook(filepath)
    ws = wb["SP500 Liquidity"]
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 2
    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"Liquidity data added to: {filepath}")
    return latest
"""
def generate_breadth_sheet(filepath, global_df, momentum_df):
    ma_df = (
        global_df
        .sort_values(["Ticker", "Date"])
        .groupby("Ticker")
        .apply(lambda g: g.assign(M20=g['Close'].rolling(20).mean()))
        .reset_index(drop=True)
    )

    latest = (
        ma_df.sort_values("Date")
        .groupby("Ticker")
        .tail(1)[["Ticker", "Close", "M20"]]
    )

    df = latest.merge(momentum_df, on ="Ticker", how="left")

    df["Above_M20"] = df["Close"] > df["M20"]
    print(df.columns.tolist())
    df["Postive Today"] = df["1-Day Return (%)"] > 0
    df["Positive Momentum"] = df["20-Day Return (%)"] > 0

    mode = "a" if os.path.exists(filepath) else "w"
    with pd.ExcelWriter(filepath, engine="openpyxl", mode=mode) as writer:
        df.to_excel(writer, sheet_name="SP500 Breadth", index=False)
    wb = load_workbook(filepath)
    ws = wb["SP500 Breadth"]
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 2
    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"Breadth data added to: {filepath}")
    return df
"""

def generate_full_corr_matrix(returns_df, output_path):
    corr = returns_df.corr()
    corr.to_csv(output_path.replace(".png", "_matrix.csv"))
    return corr


def generate_intrasector_corr_matrices(returns_df, sector_map, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    common = returns_df.columns.intersection(sector_map.index)
    returns_df = returns_df[common]
    sector_map = sector_map.loc[common]

    matrix_paths = []

    for sector in sorted(sector_map.unique()):
        tickers = sector_map[sector_map == sector].index
        sub_returns = returns_df[tickers]

        if sub_returns.shape[1] < 3:
            print(f"Skipping sector {sector} -- too few tickers.")
            continue

        corr = sub_returns.corr().replace([np.inf, -np.inf], np.nan).fillna(0)

        # Save raw matrix
        raw_path = os.path.join(output_dir, f"{sector}_corr_matrix.csv")
        corr.to_csv(raw_path)

        # Compute clustered ordering
        grid = sns.clustermap(corr, cmap="viridis")
        row_order = grid.dendrogram_row.reordered_ind
        col_order = grid.dendrogram_col.reordered_ind
        plt.close()

        clustered_corr = corr.iloc[row_order, col_order]

        # Save clustered matrix
        clustered_path = os.path.join(output_dir, f"{sector}_clustered_corr_matrix.csv")
        clustered_corr.to_csv(clustered_path)

        matrix_paths.append(clustered_path)

    return matrix_paths

def generate_sector_avg_heatmap(returns_df, sector_map, output_path):
       # Align returns_df to sector_map
    common = returns_df.columns.intersection(sector_map.index)
    returns_df = returns_df[common]
    sector_map = sector_map.loc[common]

    # Normalize sector names
    sector_map = sector_map.str.strip().str.title()

    # Drop tickers with zero variance
    returns_df = returns_df.loc[:, returns_df.std() > 0]

    # Group tickers by sector and compute sector-level average returns
    sector_returns = returns_df.groupby(sector_map, axis=1).mean()

    # Compute sector-level correlation matrix
    sector_corr = sector_returns.corr()

    mask = np.eye(sector_corr.shape[0], dtype = bool)
    sns.heatmap(
        sector_corr,
        mask = mask,
        cmap = "RdYlGn",
        vmin = 0.0,
        vmax = 1.0,
        center = 0,
        annot = True,
        fmt = ".2f",
        square = True,
        linewidths = 0.5,
        cbar_kws = {"shrink": 0.75}
    )
    plt.gcf().set_size_inches(12, 10)
    plt.tight_layout()
    plt.title("S&P 500 Sector Average Correlation Heatmap", fontsize = 18, pad = 20)
    plt.savefig(output_path, dpi = 300)
    plt.close()

    return sector_corr

def generate_full_corr_heatmap(returns_df, output_path):
    corr = returns_df.corr()
    csv_path = output_path.replace(".png", "_matrix.csv")
    corr.to_csv(csv_path)
    cppExportPath = r"D:\VB_learning\Reports\CSV_Exports\full_index_corr_matrix.csv"
    corr.to_csv(cppExportPath)
    plt.figure(figsize=(20, 16))
    sns.heatmap(
        corr,
        cmap="coolwarm",
        center=0,
        vmin=-1,
        vmax=1,
        square=True,
        cbar_kws={"shrink": 0.5},
        xticklabels=False,
        yticklabels=False,
        linewidths=0.0
    )
    plt.title("S&P 500 Full Correlation Heatmap", fontsize=20, pad=20)
    plt.tight_layout()
    plt.savefig(output_path, dpi=300)
    plt.close()

    return corr

def build_output_filepath():
    output_dir = r"D:\VB_learning\reports"
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%m-%d-%Y_%H-%M-%S")
    filename = f"sp500_report_{timestamp}.xlsx"

    return os.path.join(output_dir, filename)

def build_heatmap_filepath(name, subfolder = None):
    output_dir = r"D:\VB_learning\reports\heatmaps"
    flag = True
    if subfolder:
        output_dir = os.path.join(output_dir, subfolder)
        flag = False
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%m-%d-%Y_%H-%M-%S")

    if flag:
        filename = f"{name}_{timestamp}.png"
    else:
        filename = f"{name}_{timestamp}"
    return os.path.join(output_dir, filename)

def build_global_price_history(period = "6mo", batch_size = 50): #make the period variable in the final build
    sp500 = pd.read_csv("https://raw.githubusercontent.com/datasets/s-and-p-500-companies/master/data/constituents.csv")
    tickers = [t.replace(".", "-") for t in sp500["Symbol"].tolist()]

    frames = []
    for i in range(0, len(tickers), batch_size):
        batch = tickers[i:i + batch_size]
        print(f"Downloading batch {i // batch_size + 1} of {(len(tickers) - 1) // batch_size + 1}")

        data = yf.download(
            batch,
            period = period,
            group_by = "ticker",
            auto_adjust = False
        )

        for ticker in batch:
            if ticker not in data: 
                continue
            
            df = data[ticker].copy()
            df["Ticker"] = ticker
            df = df.reset_index()
            frames.append(df)
    global_df = pd.concat(frames, ignore_index = True)

    global_df["Daily Returns"] = (
        global_df.groupby("Ticker")["Close"].pct_change() * 100
    )

    global_df = global_df.sort_values(["Ticker", "Date"])

    return global_df # might be useful to save this dataframe as a csv for ML models later

def export_xlsx_to_csvs(xlsx_path, output_dir):
    os.makedirs(output_dir, exist_ok = True)
    xls = pd.ExcelFile(xlsx_path)
    csv_path = {}
    for sheet in xls.sheet_names:
        df = pd.read_excel(xlsx_path, sheet_name = sheet)
        safe_name = sheet.strip().replace(" ", "_").lower()
        csv_path[sheet] = os.path.join(output_dir, f"{safe_name}.csv")
        df.to_csv(csv_path[sheet], index=False)
    return csv_path

def main():
    filepath = build_output_filepath()

    global_df = build_global_price_history(period = "6mo", batch_size = 50)

    cluster_path = build_heatmap_filepath("clustered_heatmap", subfolder="intrasector")
    sector_path  = build_heatmap_filepath("sector_avg_heatmap")
    full_path    = build_heatmap_filepath("full_corr_heatmap")

    index_df, constituents_df, vol_df, daily_returns, sector_df, mom_df, sector_map = None, None, None, None, None, None, None

    if args.index:
        index_df = generate_index_sheet(filepath)
       
    if args.constituents:
        constituents_df = generate_constituent_sheet(filepath, global_df)

    if args.volatility:
        vol_df, daily_returns = generate_volatility_sheet(filepath, global_df)

    if args.momentum:
        mom_df = generate_momentum_sheet(filepath, global_df)
    
    time.sleep(0.25)

    if args.sector:
        (sector_df, sector_map) = generate_sector_sheet(filepath)
    
    if args.risk:
        corr = generate_full_corr_matrix(daily_returns, full_path)
        generate_sector_avg_heatmap(daily_returns, sector_map, sector_path)
        intrasector_paths = generate_intrasector_corr_matrices(daily_returns, sector_map, cluster_path)
        generate_risk_sheet(filepath, corr, sector_df, vol_df, mom_df)

    if args.liquidity:
        generate_liquidity_sheet(filepath, global_df, sector_df, sector_map)

   # if args.breadth:
   #     generate_breadth_sheet(filepath, global_df, mom_df)

    export_xlsx_to_csvs(filepath, output_dir = r"D:\VB_learning\reports\CSV_Exports")
    print(f"Report saved to: {filepath}")
if __name__ == "__main__":
    main()